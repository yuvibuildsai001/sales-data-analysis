"""
report_generator.py
-------------------
Creates a formatted multi-sheet Excel report from the analysis results.
Saved to reports/Sales_Analysis_Report.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, LineChart, Reference
import os
from datetime import datetime


# ── Style constants ────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1E40AF")   # Dark blue
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL  = PatternFill("solid", fgColor="EFF6FF")   # Light blue
BORDER_SIDE   = Side(style="thin", color="CBD5E1")
THIN_BORDER   = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE,
)
TITLE_FONT    = Font(bold=True, size=14, color="1E3A8A")


def _style_header_row(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_dataframe(ws, df: pd.DataFrame, start_row: int = 2, start_col: int = 1) -> None:
    """Write a DataFrame to worksheet with styled headers and alternating rows."""
    # Header
    for col_idx, col_name in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=col_idx, value=col_name.replace("_", " ").title())
    _style_header_row(ws, start_row, len(df.columns))

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

    # Auto-fit column widths
    for col_idx, col_name in enumerate(df.columns, start=start_col):
        max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)


# ── Sheet 1: Executive Summary ─────────────────────────────────────────────────
def _create_summary_sheet(wb: Workbook, df: pd.DataFrame, sql_results: dict) -> None:
    ws = wb.active
    ws.title = "Executive Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    ws["A1"] = "📊  E-Commerce Sales Analysis Report — 2024"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated on: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws["A2"].font = Font(italic=True, color="64748B", size=10)

    kpis = [
        ("Total Revenue",        f"₹{df['total_amount'].sum():,.0f}"),
        ("Total Orders",         f"{df['order_id'].nunique():,}"),
        ("Avg Order Value",      f"₹{df['total_amount'].mean():,.0f}"),
        ("Top Category",         sql_results['category_performance'].iloc[0]['category']),
        ("Top City",             sql_results['top_cities'].iloc[0]['customer_city']),
        ("Most Used Payment",    sql_results['payment_distribution'].iloc[0]['payment_method']),
        ("Avg Customer Rating",  f"{df['rating'].mean():.2f} / 5.00"),
        ("Total Products",       str(df['product_name'].nunique())),
    ]

    ws["A4"] = "KEY PERFORMANCE INDICATORS"
    ws["A4"].font = Font(bold=True, size=12, color="1E3A8A")

    for i, (label, value) in enumerate(kpis, start=5):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=i, column=2, value=value)
        cell.fill = ALT_ROW_FILL if i % 2 == 0 else PatternFill()
        cell.alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=1).border = THIN_BORDER
        cell.border = THIN_BORDER


# ── Sheet 2–5: SQL result tables ───────────────────────────────────────────────
def _create_data_sheets(wb: Workbook, sql_results: dict) -> None:
    sheets = [
        ("Monthly Revenue",   "monthly_revenue"),
        ("Category Analysis", "category_performance"),
        ("Top Cities",        "top_cities"),
        ("Payment Methods",   "payment_distribution"),
        ("Top Products",      "top_products"),
        ("Quarterly Summary", "quarterly_summary"),
    ]
    for sheet_name, key in sheets:
        ws = wb.create_sheet(title=sheet_name)
        ws["A1"] = sheet_name
        ws["A1"].font = TITLE_FONT
        _write_dataframe(ws, sql_results[key], start_row=3)


# ── Master function ────────────────────────────────────────────────────────────
def generate_excel_report(df: pd.DataFrame, sql_results: dict,
                           output_path: str = "reports/Sales_Analysis_Report.xlsx") -> str:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    _create_summary_sheet(wb, df, sql_results)
    _create_data_sheets(wb, sql_results)

    wb.save(output_path)
    print(f"📄  Excel report saved  →  {output_path}")
    return output_path
